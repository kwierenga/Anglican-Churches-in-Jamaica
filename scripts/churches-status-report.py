"""Generate an Excel report of churches by parish, showing which have
narratives, references, and pictures.

Reads:
  data/churches.csv          — canonical church list
  content/churches/*.md      — narrative markdown (presence + ## References)
  data/media.csv             — picture rows (counts per church_id)

Writes:
  churches-status.xlsx       — summary + per-parish detail sheets
"""
import csv
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
CHURCHES = ROOT / "data" / "churches.csv"
MEDIA = ROOT / "data" / "media.csv"
NARR_DIR = ROOT / "content" / "churches"
OUT = ROOT / "churches-status.xlsx"

PARISH_ORDER = [
    "Kingston", "St. Andrew", "St. Thomas", "Portland", "St. Mary", "St. Ann",
    "Trelawny", "St. James", "Hanover", "Westmoreland", "St. Elizabeth",
    "Manchester", "Clarendon", "St. Catherine",
]

# --- load data ---------------------------------------------------------------

churches = []
with CHURCHES.open(encoding="utf-8", newline="") as f:
    for row in csv.DictReader(f):
        churches.append(row)

media_count = defaultdict(int)
with MEDIA.open(encoding="utf-8", newline="") as f:
    for row in csv.DictReader(f):
        cid = row.get("church_id", "")
        if cid:
            media_count[cid] += 1

def narrative_state(cid: str) -> tuple[bool, bool]:
    p = NARR_DIR / f"{cid}.md"
    if not p.exists():
        return False, False
    text = p.read_text(encoding="utf-8")
    return True, "## References" in text

# --- enrich ------------------------------------------------------------------

for c in churches:
    has_narr, has_refs = narrative_state(c["id"])
    c["has_narrative"] = has_narr
    c["has_references"] = has_refs
    c["picture_count"] = media_count.get(c["id"], 0)
    c["has_pictures"] = c["picture_count"] > 0

# --- build workbook ----------------------------------------------------------

wb = Workbook()
ws_sum = wb.active
ws_sum.title = "Summary"

bold = Font(bold=True)
header_fill = PatternFill(start_color="1E2D4E", end_color="1E2D4E", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center")
yes_fill = PatternFill(start_color="DCEAD2", end_color="DCEAD2", fill_type="solid")
no_fill = PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")

# --- summary sheet -----------------------------------------------------------

headers = [
    "Parish", "Total churches",
    "Narratives", "Missing narrative",
    "References", "Missing references",
    "With pictures", "Without pictures",
    "Coverage (pic %)",
]
for col, h in enumerate(headers, 1):
    cell = ws_sum.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

by_parish = defaultdict(list)
for c in churches:
    by_parish[c["parish"]].append(c)

# Order: known parishes first, then any others alphabetically
parishes_seen = list(by_parish.keys())
ordered = [p for p in PARISH_ORDER if p in by_parish] + sorted(
    p for p in parishes_seen if p not in PARISH_ORDER
)

totals = dict(total=0, narr=0, refs=0, pics=0)
for r, parish in enumerate(ordered, 2):
    rows = by_parish[parish]
    total = len(rows)
    narr = sum(1 for c in rows if c["has_narrative"])
    refs = sum(1 for c in rows if c["has_references"])
    pics = sum(1 for c in rows if c["has_pictures"])
    pct = f"{pics * 100 / total:.0f}%" if total else "—"
    ws_sum.cell(row=r, column=1, value=parish)
    ws_sum.cell(row=r, column=2, value=total)
    ws_sum.cell(row=r, column=3, value=narr)
    ws_sum.cell(row=r, column=4, value=total - narr)
    ws_sum.cell(row=r, column=5, value=refs)
    ws_sum.cell(row=r, column=6, value=total - refs)
    ws_sum.cell(row=r, column=7, value=pics)
    ws_sum.cell(row=r, column=8, value=total - pics)
    ws_sum.cell(row=r, column=9, value=pct).alignment = center
    totals["total"] += total
    totals["narr"] += narr
    totals["refs"] += refs
    totals["pics"] += pics

last = len(ordered) + 2
ws_sum.cell(row=last, column=1, value="TOTAL").font = bold
ws_sum.cell(row=last, column=2, value=totals["total"]).font = bold
ws_sum.cell(row=last, column=3, value=totals["narr"]).font = bold
ws_sum.cell(row=last, column=4, value=totals["total"] - totals["narr"]).font = bold
ws_sum.cell(row=last, column=5, value=totals["refs"]).font = bold
ws_sum.cell(row=last, column=6, value=totals["total"] - totals["refs"]).font = bold
ws_sum.cell(row=last, column=7, value=totals["pics"]).font = bold
ws_sum.cell(row=last, column=8, value=totals["total"] - totals["pics"]).font = bold
ws_sum.cell(
    row=last, column=9,
    value=f"{totals['pics'] * 100 / totals['total']:.0f}%" if totals["total"] else "—",
).font = bold

for col in range(1, len(headers) + 1):
    ws_sum.column_dimensions[get_column_letter(col)].width = 18
ws_sum.column_dimensions["A"].width = 16
ws_sum.freeze_panes = "A2"

# --- detail sheet (all churches, one row each) -------------------------------

ws_det = wb.create_sheet("All churches")
det_headers = [
    "Parish", "Town", "Church", "Class", "Status",
    "Narrative", "References", "Pictures (count)", "ID",
]
for col, h in enumerate(det_headers, 1):
    cell = ws_det.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

r = 2
for parish in ordered:
    rows = sorted(by_parish[parish], key=lambda c: (c["town"] or "", c["name"]))
    for c in rows:
        ws_det.cell(row=r, column=1, value=parish)
        ws_det.cell(row=r, column=2, value=c["town"])
        ws_det.cell(row=r, column=3, value=c["name"])
        ws_det.cell(row=r, column=4, value=c["classification"])
        ws_det.cell(row=r, column=5, value=c["status"])
        narr = ws_det.cell(row=r, column=6, value="Yes" if c["has_narrative"] else "—")
        narr.fill = yes_fill if c["has_narrative"] else no_fill
        narr.alignment = center
        refs = ws_det.cell(row=r, column=7, value="Yes" if c["has_references"] else "—")
        refs.fill = yes_fill if c["has_references"] else no_fill
        refs.alignment = center
        pic_val = c["picture_count"] if c["has_pictures"] else "—"
        pic = ws_det.cell(row=r, column=8, value=pic_val)
        pic.fill = yes_fill if c["has_pictures"] else no_fill
        pic.alignment = center
        ws_det.cell(row=r, column=9, value=c["id"])
        r += 1

widths = [16, 22, 36, 14, 10, 12, 12, 16, 48]
for col, w in enumerate(widths, 1):
    ws_det.column_dimensions[get_column_letter(col)].width = w
ws_det.freeze_panes = "A2"
ws_det.auto_filter.ref = f"A1:{get_column_letter(len(det_headers))}{r-1}"

# --- gaps sheet (only churches missing something) ----------------------------

ws_gap = wb.create_sheet("Gaps")
gap_headers = [
    "Parish", "Town", "Church", "Class", "Status",
    "Missing narrative", "Missing references", "Missing pictures", "ID",
]
for col, h in enumerate(gap_headers, 1):
    cell = ws_gap.cell(row=1, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center

r = 2
for parish in ordered:
    rows = sorted(by_parish[parish], key=lambda c: (c["town"] or "", c["name"]))
    for c in rows:
        if c["has_narrative"] and c["has_references"] and c["has_pictures"]:
            continue
        ws_gap.cell(row=r, column=1, value=parish)
        ws_gap.cell(row=r, column=2, value=c["town"])
        ws_gap.cell(row=r, column=3, value=c["name"])
        ws_gap.cell(row=r, column=4, value=c["classification"])
        ws_gap.cell(row=r, column=5, value=c["status"])
        for col, key in (
            (6, "has_narrative"), (7, "has_references"), (8, "has_pictures"),
        ):
            mc = ws_gap.cell(row=r, column=col, value="" if c[key] else "missing")
            if not c[key]:
                mc.fill = no_fill
                mc.alignment = center
        ws_gap.cell(row=r, column=9, value=c["id"])
        r += 1

for col, w in enumerate(widths[:-1] + [48], 1):
    ws_gap.column_dimensions[get_column_letter(col)].width = w
ws_gap.freeze_panes = "A2"
if r > 2:
    ws_gap.auto_filter.ref = f"A1:{get_column_letter(len(gap_headers))}{r-1}"

wb.save(OUT)

print(f"Wrote {OUT}")
print(f"  Churches: {len(churches)}")
print(f"  With narrative:  {totals['narr']}/{totals['total']}")
print(f"  With references: {totals['refs']}/{totals['total']}")
print(f"  With pictures:   {totals['pics']}/{totals['total']}")
