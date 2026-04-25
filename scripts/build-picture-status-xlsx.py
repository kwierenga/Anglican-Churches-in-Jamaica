#!/usr/bin/env python
"""Generate churches-picture-status.xlsx from churches.csv and media.csv.

Sheet layout:
- "Summary": one row per parish with count/total.
- One sheet per parish: rows sorted by town, columns Town / Church / Class / Status / Pics.
"""
import csv
import os
from collections import defaultdict

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CHURCHES = os.path.join(ROOT, "data/churches.csv")
MEDIA = os.path.join(ROOT, "data/media.csv")
OUT = os.path.join(ROOT, "churches-picture-status.xlsx")

img_count = defaultdict(int)
with open(MEDIA, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        cid = (row.get("church_id") or "").strip()
        if cid:
            img_count[cid] += 1

parishes = defaultdict(list)
with open(CHURCHES, encoding="utf-8") as f:
    for row in csv.DictReader(f):
        parishes[row["parish"]].append(row)

wb = Workbook()

HEADER_FILL = PatternFill("solid", fgColor="1E2D4E")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HAS_PICS_FILL = PatternFill("solid", fgColor="E8F5E9")
NO_PICS_FILL = PatternFill("solid", fgColor="FFEBEE")
INACTIVE_FONT = Font(italic=True, color="808080")
BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

def style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

# --- Summary sheet ---
summary = wb.active
summary.title = "Summary"
summary.append(["Parish", "Churches", "With pictures", "Coverage", "Total images"])
style_header(summary, 5)

grand_total = 0
grand_with = 0
grand_imgs = 0
for parish in sorted(parishes.keys()):
    entries = parishes[parish]
    with_p = sum(1 for e in entries if img_count.get(e["id"], 0) > 0)
    total_imgs = sum(img_count.get(e["id"], 0) for e in entries)
    coverage = with_p / len(entries) if entries else 0
    summary.append([parish, len(entries), with_p, coverage, total_imgs])
    grand_total += len(entries)
    grand_with += with_p
    grand_imgs += total_imgs

summary.append([
    "TOTAL", grand_total, grand_with,
    grand_with / grand_total if grand_total else 0, grand_imgs,
])
total_row = summary.max_row
for c in range(1, 6):
    summary.cell(row=total_row, column=c).font = Font(bold=True)

for i in range(2, summary.max_row + 1):
    summary.cell(row=i, column=4).number_format = "0%"

summary.column_dimensions["A"].width = 18
summary.column_dimensions["B"].width = 12
summary.column_dimensions["C"].width = 16
summary.column_dimensions["D"].width = 11
summary.column_dimensions["E"].width = 14

# --- Per-parish sheets ---
for parish in sorted(parishes.keys()):
    entries = sorted(parishes[parish], key=lambda r: r["town"].lower())
    with_p = sum(1 for e in entries if img_count.get(e["id"], 0) > 0)
    ws = wb.create_sheet(title=parish[:31])  # Excel sheet name limit
    ws.append(["Town", "Church", "Class", "Status", "Pics"])
    style_header(ws, 5)
    for e in entries:
        n = img_count.get(e["id"], 0)
        ws.append([e["town"], e["name"], e["classification"], e["status"], n])
        r = ws.max_row
        fill = HAS_PICS_FILL if n > 0 else NO_PICS_FILL
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = fill
            ws.cell(row=r, column=c).border = BORDER
            if e["status"] in ("inactive", "ruin"):
                ws.cell(row=r, column=c).font = INACTIVE_FONT
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 8
    ws.auto_filter.ref = f"A1:E{ws.max_row}"
    # title row: parish + coverage — inject via sheet name only; caption is optional

wb.save(OUT)
print(f"wrote {OUT}")
print(f"  total churches: {grand_total}")
print(f"  with pictures:  {grand_with}")
print(f"  total images:   {grand_imgs}")
